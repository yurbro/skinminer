from __future__ import annotations

import ast
import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "human_annotation"

SELECTED_PAPERS = [
    {
        "index": 1,
        "doi": "10.1208/s12249-013-9995-4",
        "difficulty": "Easy",
        "route": "table",
        "selection_reason": "Core positive benchmark: OA HTML table, 24 assembled rows, 24 verified under extended policy, and fully covered in Round 2 gold.",
    },
    {
        "index": 2,
        "doi": "10.1039/d0ra00100g",
        "difficulty": "Easy",
        "route": "table",
        "selection_reason": "Clear OA table paper with structured 24 h endpoint rows. Useful as an easy negative-control table case from Round 2 gold.",
    },
    {
        "index": 3,
        "doi": "10.1186/2050-6511-13-5",
        "difficulty": "Medium",
        "route": "mixed",
        "selection_reason": "OA mixed-route paper with substantial assembled output and unresolved scope/concentration issues. Human extraction also requires cross-checking text and figures.",
    },
    {
        "index": 4,
        "doi": "10.1007/s11095-008-9785-y",
        "difficulty": "Medium",
        "route": "mixed",
        "selection_reason": "OA mixed-route scope-confusion case from Round 2 gold. Good for testing whether humans also separate IVPT-like wording from in vivo/DPK context.",
    },
    {
        "index": 5,
        "doi": "10.1016/j.ejpb.2020.05.013",
        "difficulty": "Hard",
        "route": "figure",
        "selection_reason": "OA figure-heavy paper with one retained positive in Round 2. Manual figure reading is required and directly comparable to SkinMiner figure extraction.",
    },
    {
        "index": 6,
        "doi": "10.1016/j.ijpharm.2016.03.043",
        "difficulty": "Hard",
        "route": "figure",
        "selection_reason": "OA figure-heavy comparison paper from Round 2 gold. Human readers must recover endpoints from multi-curve plots across several formulations.",
    },
]

TEMPLATE_COLUMNS = [
    "start_time",
    "end_time",
    "formulation_label",
    "api_name",
    "api_concentration",
    "api_concentration_unit",
    "endpoint_type",
    "endpoint_value",
    "endpoint_unit",
    "endpoint_time",
    "endpoint_time_unit",
    "device",
    "membrane_type",
    "membrane_source",
    "receptor_medium",
    "dose_type",
    "dose_amount",
    "source_evidence",
    "confidence",
    "notes",
]


def ensure_output_dir() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def parse_literal(value: object, default: object) -> object:
    if not isinstance(value, str) or not value.strip():
        return default
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        try:
            return ast.literal_eval(value)
        except (ValueError, SyntaxError):
            return default


def load_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    assembled = pd.read_csv(ROOT / "outputs" / "full_run_16_post_all_fixes" / "assembled_records.csv")
    v2 = pd.read_csv(ROOT / "outputs" / "full_run_16_post_all_fixes" / "v2_rescore" / "verified_records.csv")
    content_access = pd.read_csv(ROOT / "outputs" / "full_run_16_post_all_fixes" / "content_access.csv")
    gold = pd.read_csv(ROOT / "outputs" / "gold_audit_set" / "round2" / "gold_set_round2_annotation.csv")
    return assembled, v2, content_access, gold


def preferred_access_url(row: pd.Series) -> str:
    access_urls = parse_literal(row.get("access_urls"), {})
    if not isinstance(access_urls, dict):
        return f"https://doi.org/{row['doi']}"
    preferred_format = row.get("preferred_format")
    if preferred_format in access_urls:
        return access_urls[preferred_format]
    if "html" in access_urls:
        return access_urls["html"]
    if "pdf" in access_urls:
        return access_urls["pdf"]
    return f"https://doi.org/{row['doi']}"


def format_source_evidence(row: pd.Series) -> str:
    evidence = parse_literal(row.get("evidence_json"), [])
    provenance = parse_literal(row.get("provenance_json"), {})
    locators: list[str] = []
    if isinstance(evidence, list):
        for item in evidence:
            if not isinstance(item, dict):
                continue
            locator = str(item.get("locator") or "").strip()
            page = item.get("page")
            if locator:
                if page not in (None, "", "nan") and "page" not in locator.lower():
                    locators.append(f"{locator} (p.{page})")
                else:
                    locators.append(locator)
    metadata = provenance.get("metadata", {}) if isinstance(provenance, dict) else {}
    for key in ["row_source_locators", "selected_locators"]:
        value = metadata.get(key)
        if isinstance(value, list):
            for locator in value:
                if locator:
                    locators.append(str(locator))
    unique_locators: list[str] = []
    for locator in locators:
        if locator not in unique_locators:
            unique_locators.append(locator)
    if unique_locators:
        return "; ".join(unique_locators[:3])
    source_path = str(row.get("source_path") or "").strip()
    return Path(source_path).name if source_path else ""


def build_selected_paper_rows(
    assembled: pd.DataFrame,
    v2: pd.DataFrame,
    content_access: pd.DataFrame,
    gold: pd.DataFrame,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for item in SELECTED_PAPERS:
        doi = item["doi"]
        assembled_sub = assembled[assembled["doi"] == doi]
        v2_sub = v2[v2["doi"] == doi]
        access_row = content_access[content_access["doi"] == doi].iloc[0]
        gold_sub = gold[gold["doi"] == doi]
        rows.append(
            {
                "index": item["index"],
                "doi": doi,
                "title": access_row["title"],
                "difficulty": item["difficulty"],
                "route": item["route"],
                "pipeline_records": int(len(assembled_sub)),
                "verified_v2": int((v2_sub["verification_status"] == "verified").sum()),
                "selection_reason": item["selection_reason"],
                "oa_access_url": preferred_access_url(access_row),
                "doi_url": f"https://doi.org/{doi}",
                "gold_rows": int(len(gold_sub)),
            }
        )
    return rows


def write_paper_selection(rows: list[dict[str, object]]) -> None:
    lines = [
        "# Human Annotation Experiment: Paper Selection",
        "",
        "## Selected Papers",
        "",
        "| # | DOI | Title | Difficulty | Route | Pipeline records | Verified (extended policy) | Selection reason |",
        "|---|---|---|---|---|---:|---:|---|",
    ]
    for row in rows:
        lines.append(
            f"| {row['index']} | {row['doi']} | {row['title']} | {row['difficulty']} | {row['route']} | {row['pipeline_records']} | {row['verified_v2']} | {row['selection_reason']} |"
        )
    lines.extend(
        [
            "",
            "Difficulty distribution is deliberate: 2 table-easy, 2 mixed-medium, 2 figure-hard. All six papers are drawn from `full_run_16_post_all_fixes` assembled records and all six have downloadable OA full text.",
            "",
            "## Paper Access Links",
            "",
        ]
    )
    for row in rows:
        lines.append(
            f"{row['index']}. `{row['doi']}`: DOI link {row['doi_url']} | OA access {row['oa_access_url']}"
        )
    (OUTPUT_DIR / "paper_selection.md").write_text("\n".join(lines), encoding="utf-8")


def style_sheet(ws, widths: dict[str, float], freeze_panes: str | None = None) -> None:
    if freeze_panes:
        ws.freeze_panes = freeze_panes
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_excel_template(rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Instructions for Manual Data Extraction"
    ws["A1"].font = Font(bold=True, size=12)
    instructions = [
        "Task: Read each paper and extract all in vitro permeation/release data records.",
        "",
        "For each experiment/formulation reported in the paper, fill in one row with:",
        "- Formulation label (e.g., F1, F2, Formulation A)",
        "- API name (active pharmaceutical ingredient)",
        "- API concentration and unit",
        "- Endpoint type (cumulative amount / flux / Jss)",
        "- Endpoint value and unit",
        "- Endpoint timepoint and unit",
        "- Device (Franz cell / other)",
        "- Membrane type and source",
        "- Receptor medium",
        "- Dose type (finite / infinite) and amount",
        "- Source evidence (which table/figure/section)",
        "",
        "Rules:",
        "1. Extract ALL formulations and ALL timepoints, not just summary values.",
        "2. If a table has multiple timepoints, create one row per formulation × timepoint.",
        "3. Only extract in vitro data (not in vivo / clinical / DPK).",
        "4. If a value is not explicitly stated, leave blank; do not guess.",
        "5. Record your start time and end time for EACH paper.",
        "",
        "Please record your experience level:",
        "- PhD student in pharmaceutical sciences",
        "- Postdoc / researcher in formulation science",
        "- Other: ______",
    ]
    for idx, line in enumerate(instructions, start=3):
        ws[f"A{idx}"] = line
    style_sheet(ws, {"A": 120})

    for row in rows:
        ws = wb.create_sheet(f"Paper {row['index']}")
        ws["A1"] = "DOI"
        ws["B1"] = row["doi"]
        ws["A2"] = "Title"
        ws["B2"] = row["title"]
        ws["A3"] = "Difficulty"
        ws["B3"] = row["difficulty"]
        ws["D1"] = "OA access"
        ws["E1"] = row["oa_access_url"]
        for header_index, header in enumerate(TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=5, column=header_index)
            cell.value = header
            cell.font = Font(bold=True)
        for _ in range(150):
            ws.append([""] * len(TEMPLATE_COLUMNS))
        style_sheet(
            ws,
            {
                "A": 11,
                "B": 11,
                "C": 20,
                "D": 18,
                "E": 15,
                "F": 18,
                "G": 16,
                "H": 14,
                "I": 14,
                "J": 12,
                "K": 14,
                "L": 18,
                "M": 20,
                "N": 18,
                "O": 22,
                "P": 14,
                "Q": 14,
                "R": 28,
                "S": 12,
                "T": 28,
            },
            freeze_panes="A6",
        )

    summary = wb.create_sheet("Summary")
    summary_headers = ["Paper", "DOI", "Start time", "End time", "Duration (min)", "Records extracted", "Difficulty rating (1-5)", "Notes"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = summary.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
    for row_idx, row in enumerate(rows, start=2):
        summary.cell(row=row_idx, column=1).value = f"Paper {row['index']}"
        summary.cell(row=row_idx, column=2).value = row["doi"]
    style_sheet(summary, {"A": 10, "B": 28, "C": 12, "D": 12, "E": 14, "F": 16, "G": 18, "H": 30}, freeze_panes="A2")

    wb.save(OUTPUT_DIR / "extraction_template.xlsx")


def write_skinminer_reference(assembled: pd.DataFrame, v2: pd.DataFrame) -> int:
    selected_dois = [row["doi"] for row in SELECTED_PAPERS]
    sub = assembled[assembled["doi"].isin(selected_dois)].copy()
    v2_status = v2[["record_id", "verification_status"]].rename(columns={"verification_status": "pipeline_verification_status"})
    sub = sub.merge(v2_status, on="record_id", how="left")
    rows = []
    for _, row in sub.iterrows():
        rows.append(
            {
                "doi": row["doi"],
                "formulation_label": row["formulation_label"],
                "api_name": row["api_name"],
                "api_concentration": row["api_concentration_value"] if pd.notna(row["api_concentration_value"]) else "",
                "api_concentration_unit": row["api_concentration_unit"] if pd.notna(row["api_concentration_unit"]) else "",
                "endpoint_type": row["endpoint_kind"] if pd.notna(row["endpoint_kind"]) else "",
                "endpoint_value": row["endpoint_value"] if pd.notna(row["endpoint_value"]) else "",
                "endpoint_unit": row["endpoint_unit"] if pd.notna(row["endpoint_unit"]) else "",
                "endpoint_time": row["endpoint_time_value"] if pd.notna(row["endpoint_time_value"]) else "",
                "endpoint_time_unit": row["endpoint_time_unit"] if pd.notna(row["endpoint_time_unit"]) else "",
                "device": row["device"] if pd.notna(row["device"]) else "",
                "membrane_type": row["barrier"] if pd.notna(row["barrier"]) else "",
                "membrane_source": "",
                "receptor_medium": "",
                "dose_type": "",
                "dose_amount": "",
                "source_evidence": format_source_evidence(row),
                "pipeline_verification_status": row["pipeline_verification_status"] if pd.notna(row["pipeline_verification_status"]) else "",
            }
        )
    out_path = OUTPUT_DIR / "skinminer_reference.csv"
    with out_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    return len(rows)


def write_evaluation_framework() -> None:
    text = """# Human vs SkinMiner Comparison Framework

## Metrics

1. **Extraction time**
   - Per-paper mean ± SD across annotators
   - Total time for 6 papers
   - SkinMiner time for the same 6 papers from pipeline logs

2. **Record coverage**
   - Human records per paper vs SkinMiner records per paper
   - Unique records found by human only / SkinMiner only / both

3. **Value accuracy**
   - For records found by both: value match rate
   - Numeric tolerance: ±5%

4. **Inter-annotator agreement**
   - Between human annotators: record-level agreement
   - Between each human annotator and SkinMiner: record-level agreement

5. **Cost comparison**
   - Human: hours × hourly rate (for example UK PhD student ~£12-15/hour)
   - SkinMiner: API cost for the same 6 papers

## Expected Result Template

| Metric | Human (mean) | SkinMiner | Ratio |
|---|---|---|---|
| Time per paper (min) | ? | ? | ?x |
| Total time (6 papers) | ? | ? | ?x |
| Records per paper | ? | ? |  |
| Value accuracy | ? | ? |  |
| Cost | ? | ? | ?x |

## Notes

- Compare against `outputs/human_annotation/skinminer_reference.csv`, which contains all assembled records for the selected six papers under the extended-policy rescore.
- Do not expose the SkinMiner reference file to annotators before manual extraction is complete.
- For difficult figure papers, record whether manual endpoint reading required ruler/zoom support or approximate visual interpolation.
"""
    (OUTPUT_DIR / "evaluation_framework.md").write_text(text, encoding="utf-8")


def main() -> None:
    ensure_output_dir()
    assembled, v2, content_access, gold = load_inputs()
    rows = build_selected_paper_rows(assembled, v2, content_access, gold)
    write_paper_selection(rows)
    write_excel_template(rows)
    write_skinminer_reference(assembled, v2)
    write_evaluation_framework()


if __name__ == "__main__":
    main()
